'''
Created on Jan 28, 2021

@author: ykm09
'''
import random
from openpyxl import Workbook

gerund=["avoid","consider","delay","deny","dislike","enjoy","escape","finish","give up","imagine","keep","mind","postpone","practice","put off","stop","suggest"]
to=["agree","choose","decide","expect","fail","hope","plan","promise","refuse","want","wish","would like"]

c=list()
for i in gerund:
    c.append((i, "ing"))
for i in to:
    c.append((i, "to"))

random.shuffle(c)
write_wb = Workbook()
write_wb.remove(write_wb["Sheet"])

sheet1 = write_wb.active
sheet1 = write_wb.create_sheet("동명사&to부정사 문제지")
sheet1.column_dimensions['A'].width = 2

for i in range(len(c)):
    sheet1.cell(i+2,2,c[i][0])

sheet2 = write_wb.create_sheet('동명사&to부정사 답지') 
sheet2.column_dimensions['A'].width = 2
for i in range(len(c)):
    sheet2.cell(i+2,2,c[i][0])
    sheet2.cell(i+2,3,c[i][1])
r=["be busy -ing","be used[accustomed]to -ing","be worth -ing","cannot help -ing  =cannot (help) but +동사원형","feel like -ing","There is no -ing  =It is impossible to~","have difficulty[trouble](in) -ing","It is no use[good] -ing","keep[stop/prevent/prohibit] A from -ing","look forward to -ing","on[upon] -ing  =as soon as+주어+동사", "spend[waste]+시간/돈+ -ing", "be used to+동사원형","used to+동사원형"]
b=["~하느라 바쁘다","~하는 데 익숙하다","~할 가치가 있다","~할 수밖에 없다","~하고 싶다","~할 수 없다","~하는 데 어려움을 겪다","~해 봐야 소용없다","A가 ~하는 것을 막다","~하기를 고대하다","~하자마자","~하느라 시간/돈을 쓰다[낭비하다]","~하는 데 사용되다","(과거에)~하곤 했다"]
q=list(zip(r, b))
random.shuffle(q)
sheet3 = write_wb.create_sheet('관용표현 문제지1')
sheet3.column_dimensions['A'].width = 2
sheet3.column_dimensions['B'].width = 40
sheet3.column_dimensions['C'].width = 29
sheet4 = write_wb.create_sheet('관용표현 답지1')
sheet4.column_dimensions['A'].width = 2
sheet4.column_dimensions['B'].width = 40
sheet4.column_dimensions['C'].width = 29
random.shuffle(q)
for i in range(len(q)):
    sheet3.cell(i+2, 2, q[i][0])
    sheet4.cell(i+2, 2, q[i][0])
    sheet4.cell(i+2, 3, q[i][1])

sheet5 = write_wb.create_sheet('관용표현 문제지2')
sheet5.column_dimensions['A'].width = 2
sheet5.column_dimensions['B'].width = 29
sheet5.column_dimensions['C'].width = 40
sheet6 = write_wb.create_sheet('관용표현 답지2')
sheet6.column_dimensions['A'].width = 2
sheet6.column_dimensions['B'].width = 29
sheet6.column_dimensions['C'].width = 40
random.shuffle(q)
for i in range(len(q)):
    sheet5.cell(i+2, 2, q[i][1])
    sheet6.cell(i+2, 2, q[i][1])
    sheet6.cell(i+2, 3, q[i][0])


write_wb.save('/Users/ykm09/Desktop/영어닷!! 도망쳐!!.xlsx')

#행 단위로 추가
# write_ws.append([1,2,3])


